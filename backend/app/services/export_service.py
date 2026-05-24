from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font

from app.core.config import EXPORTS_DIR
from app.db.database import fetch_all, fetch_one
from app.services.grading_task_service import get_grading_task

REVIEW_DONE = "已复核"
OBSOLETE_DUPLICATE = "重复作废"


def _loads_json(value: str | None, default: Any) -> Any:
    if not value:
        return default
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return default


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_name(value: str | None) -> str:
    text = (value or "").strip()
    for char in '\\/:*?"<>|':
        text = text.replace(char, "_")
    return text or "未命名"


def _task_students(task: dict) -> list[dict]:
    rows = fetch_all(
        """
        SELECT s.*, tc.class_name AS teaching_class_name
        FROM students s
        LEFT JOIN teaching_classes tc ON tc.id = s.class_id
        WHERE s.course_id = ? AND s.class_id = ?
        ORDER BY s.anonymous_id ASC, s.student_no ASC, s.id ASC
        """,
        (task["course_id"], task["class_id"]),
    )
    if rows:
        return rows
    return fetch_all(
        """
        SELECT cs.*, tc.class_name AS teaching_class_name
        FROM class_students cs
        LEFT JOIN teaching_classes tc ON tc.id = cs.class_id
        WHERE cs.class_id = ?
        ORDER BY cs.anonymous_id ASC, cs.student_no ASC, cs.id ASC
        """,
        (task["class_id"],),
    )


def _task_submissions(task_id: int) -> list[dict]:
    return fetch_all(
        """
        SELECT
            s.*,
            COALESCE(s.teacher_total_score, ts.total_score) AS final_teacher_total_score,
            COALESCE(s.teacher_overall_comment, ts.feedback) AS final_teacher_comment,
            ts.item_scores_json,
            ts.reviewer_name,
            ts.confirmed_at AS teacher_confirmed_at
        FROM submissions s
        LEFT JOIN teacher_scores ts ON ts.submission_id = s.id
        WHERE COALESCE(s.task_id, s.grading_task_id) = ?
          AND COALESCE(s.match_status, '') != ?
        ORDER BY s.student_no ASC, s.id ASC
        """,
        (task_id, OBSOLETE_DUPLICATE),
    )


def _rubric_total(task_id: int) -> float:
    row = fetch_one(
        """
        SELECT SUM(COALESCE(ri.max_score, 0)) AS total
        FROM grading_tasks gt
        JOIN rubric_items ri ON ri.experiment_id = gt.experiment_id
        WHERE gt.id = ?
        """,
        (task_id,),
    )
    return round(_to_float(row.get("total") if row else 0), 2)


def _pick_submission(submissions: list[dict]) -> dict | None:
    if not submissions:
        return None
    return sorted(
        submissions,
        key=lambda item: (
            0 if item.get("review_status") == REVIEW_DONE else 1,
            0 if item.get("final_teacher_total_score") is not None else 1,
            -int(item.get("id") or 0),
        ),
    )[0]


def build_export_data(task_id: int, mode: str = "reviewed_only") -> dict:
    task = get_grading_task(task_id)
    if not task:
        raise ValueError("未找到批改任务。")

    students = _task_students(task)
    submissions = _task_submissions(task_id)
    by_student: dict[int, list[dict]] = {}
    unmatched_rows: list[dict] = []
    for submission in submissions:
        student_id = submission.get("student_id")
        if student_id is None:
            unmatched_rows.append(submission)
        else:
            by_student.setdefault(int(student_id), []).append(submission)

    grade_rows: list[dict] = []
    item_rows: list[dict] = []
    unfinished_rows: list[dict] = []
    reviewed_scores: list[float] = []
    ai_scores: list[float] = []
    item_score_values: dict[str, list[float]] = {}
    item_max_scores: dict[str, float] = {}
    item_modified_counts: dict[str, int] = {}

    for student in students:
        student_id = int(student["id"])
        submission = _pick_submission(by_student.get(student_id, []))
        review_status = (submission or {}).get("review_status") or "未提交"
        match_status = (submission or {}).get("match_status") or "未提交"
        export_status = "已复核，可导出" if review_status == REVIEW_DONE else "未复核，不建议导出"
        if submission is None:
            export_status = "未提交"
        elif match_status not in ("已匹配", "姓名不一致"):
            export_status = match_status

        if mode == "reviewed_only" and review_status != REVIEW_DONE:
            unfinished_rows.append(
                {
                    "学号": student.get("student_no") or "",
                    "姓名": student.get("student_name") or student.get("name") or "",
                    "班级": student.get("class_name") or student.get("teaching_class_name") or "",
                    "状态": export_status,
                    "提交文件名": (submission or {}).get("original_filename") or "",
                }
            )
            continue

        teacher_total = (submission or {}).get("final_teacher_total_score")
        ai_total = (submission or {}).get("ai_total_score")
        row = {
            "学号": student.get("student_no") or "",
            "姓名": student.get("student_name") or student.get("name") or "",
            "班级": student.get("class_name") or student.get("teaching_class_name") or task.get("class_name") or "",
            "匿名编号": student.get("anonymous_id") or "",
            "课程": task.get("course_name") or "",
            "教学班": task.get("class_name") or "",
            "实验任务": task.get("experiment_name") or "",
            "AI建议总分": round(_to_float(ai_total), 2) if ai_total is not None else "",
            "教师确认总分": round(_to_float(teacher_total), 2) if teacher_total is not None else "",
            "复核状态": review_status,
            "导出状态": export_status,
            "教师总评": (submission or {}).get("final_teacher_comment") or "",
            "提交文件名": (submission or {}).get("original_filename") or "",
            "匹配状态": match_status,
            "复核时间": (submission or {}).get("reviewed_at") or (submission or {}).get("teacher_confirmed_at") or "",
        }
        grade_rows.append(row)

        if review_status == REVIEW_DONE and teacher_total is not None:
            reviewed_scores.append(_to_float(teacher_total))
            if ai_total is not None:
                ai_scores.append(_to_float(ai_total))

        teacher_items = _loads_json((submission or {}).get("item_scores_json"), [])
        item_row = {"学号": row["学号"], "姓名": row["姓名"], "班级": row["班级"], "匿名编号": row["匿名编号"], "复核状态": review_status}
        if isinstance(teacher_items, list):
            for item in teacher_items:
                if not isinstance(item, dict):
                    continue
                item_name = item.get("item_name") or f"评分项{item.get('rubric_item_id') or ''}"
                score_value = round(_to_float(item.get("teacher_score")), 2)
                ai_value = round(_to_float(item.get("ai_score")), 2)
                max_value = round(_to_float(item.get("max_score")), 2)
                item_row[item_name] = score_value
                item_row[f"{item_name}备注"] = item.get("teacher_comment") or ""
                item_row[f"{item_name}AI建议分"] = ai_value
                if review_status == REVIEW_DONE:
                    item_score_values.setdefault(item_name, []).append(score_value)
                    item_max_scores[item_name] = max_value
                    if bool(item.get("is_modified")) or abs(score_value - ai_value) > 0.001:
                        item_modified_counts[item_name] = item_modified_counts.get(item_name, 0) + 1
        if mode != "reviewed_only" or review_status == REVIEW_DONE:
            item_rows.append(item_row)

        if export_status != "已复核，可导出":
            unfinished_rows.append({"学号": row["学号"], "姓名": row["姓名"], "班级": row["班级"], "状态": export_status, "提交文件名": row["提交文件名"]})

    for submission in unmatched_rows:
        unfinished_rows.append(
            {
                "学号": submission.get("parsed_student_no") or submission.get("student_no") or "",
                "姓名": submission.get("parsed_student_name") or submission.get("student_name") or "",
                "班级": task.get("class_name") or "",
                "状态": submission.get("match_status") or "未匹配",
                "提交文件名": submission.get("original_filename") or "",
            }
        )

    total_max_score = _rubric_total(task_id)
    reviewed_count = len(reviewed_scores)
    pass_line = total_max_score * 0.6 if total_max_score else 60
    summary = {
        "班级名单人数": len(students),
        "成绩记录数": len(grade_rows),
        "已复核人数": reviewed_count,
        "未完成记录数": len(unfinished_rows),
        "满分": total_max_score,
        "平均分": round(sum(reviewed_scores) / reviewed_count, 2) if reviewed_count else None,
        "最高分": round(max(reviewed_scores), 2) if reviewed_scores else None,
        "最低分": round(min(reviewed_scores), 2) if reviewed_scores else None,
        "中位数": round(float(pd.Series(reviewed_scores).median()), 2) if reviewed_scores else None,
        "及格率": round(sum(1 for score in reviewed_scores if score >= pass_line) / reviewed_count * 100, 2) if reviewed_count else None,
        "人与AI平均差值": round(sum((t - a) for t, a in zip(reviewed_scores, ai_scores)) / len(ai_scores), 2) if ai_scores else None,
    }

    distribution = []
    for label, low, high in [("90%-100%", 0.9, 1.01), ("80%-89%", 0.8, 0.9), ("70%-79%", 0.7, 0.8), ("60%-69%", 0.6, 0.7), ("60%以下", -1, 0.6)]:
        count = 0
        for score in reviewed_scores:
            ratio = score / total_max_score if total_max_score else score / 100
            if low <= ratio < high:
                count += 1
        distribution.append({"分数段": label, "人数": count})

    rubric_analysis = []
    for item_name, values in item_score_values.items():
        max_score = item_max_scores.get(item_name, 0)
        average = sum(values) / len(values) if values else 0
        rubric_analysis.append(
            {
                "评分项": item_name,
                "满分": round(max_score, 2),
                "平均分": round(average, 2),
                "平均得分率": round(average / max_score * 100, 2) if max_score else None,
                "教师改动次数": item_modified_counts.get(item_name, 0),
            }
        )

    pass_count = sum(1 for score in reviewed_scores if score >= pass_line)
    pass_fail = [{"类别": "及格", "人数": pass_count}, {"类别": "未及格", "人数": reviewed_count - pass_count}]
    ai_diff = []
    for row in grade_rows:
        teacher_total = row.get("教师确认总分")
        ai_total = row.get("AI建议总分")
        if teacher_total != "" and ai_total != "":
            ai_diff.append({"学生": f"{row['学号']} {row['姓名']}", "人与AI评分差值": round(_to_float(teacher_total) - _to_float(ai_total), 2)})

    return {
        "task": task,
        "summary": summary,
        "grades": grade_rows,
        "items": item_rows,
        "unfinished": unfinished_rows,
        "distribution": distribution,
        "rubric_analysis": rubric_analysis,
        "pass_fail": pass_fail,
        "ai_diff": ai_diff,
    }


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, rows: list[dict] | dict) -> None:
    if isinstance(rows, dict):
        df = pd.DataFrame([{"指标": key, "数值": "" if value is None else value} for key, value in rows.items()])
    else:
        df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame({"提示": ["暂无数据"]})
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        width = min(max(max(len(str(cell.value or "")) for cell in col) + 2, 12), 38)
        ws.column_dimensions[col[0].column_letter].width = width


def export_grades_to_excel(task_id: int, mode: str = "reviewed_only") -> Path:
    data = build_export_data(task_id, mode)
    task = data["task"]
    filename = (
        f"{_safe_name(task.get('course_name'))}_{_safe_name(task.get('class_name'))}_"
        f"{_safe_name(task.get('experiment_name'))}_成绩导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    path = EXPORTS_DIR / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_sheet(writer, "成绩汇总", data["grades"])
        _write_sheet(writer, "分项成绩", data["items"])
        _write_sheet(writer, "未完成名单", data["unfinished"])
        _write_sheet(writer, "班级分析", data["summary"])
        _write_sheet(writer, "分数段分布", data["distribution"])
        _write_sheet(writer, "评分项分析", data["rubric_analysis"])
    return path


def export_analysis_to_excel(task_id: int, mode: str = "reviewed_only") -> Path:
    data = build_export_data(task_id, mode)
    task = data["task"]
    filename = (
        f"{_safe_name(task.get('course_name'))}_{_safe_name(task.get('class_name'))}_"
        f"{_safe_name(task.get('experiment_name'))}_班级分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    path = EXPORTS_DIR / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_sheet(writer, "分析概览", data["summary"])
        _write_sheet(writer, "分数段分布", data["distribution"])
        _write_sheet(writer, "评分项分析", data["rubric_analysis"])
        _write_sheet(writer, "及格情况", data["pass_fail"])
        _write_sheet(writer, "人与AI差值", data["ai_diff"])
        _write_sheet(writer, "成绩明细", data["grades"])

        book = writer.book
        chart_sheet = book.create_sheet("图形分析")
        if data["distribution"]:
            chart = BarChart()
            chart.title = "分数段分布"
            chart.y_axis.title = "人数"
            chart.x_axis.title = "分数段"
            ws = writer.sheets["分数段分布"]
            chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
            chart.height = 8
            chart.width = 14
            chart_sheet.add_chart(chart, "A1")
        if data["rubric_analysis"]:
            chart = BarChart()
            chart.title = "评分项平均得分率"
            chart.y_axis.title = "平均得分率（%）"
            ws = writer.sheets["评分项分析"]
            chart.add_data(Reference(ws, min_col=4, min_row=1, max_row=ws.max_row), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
            chart.height = 8
            chart.width = 18
            chart_sheet.add_chart(chart, "A18")
        if sum(item["人数"] for item in data["pass_fail"]) > 0:
            pie = PieChart()
            pie.title = "及格情况"
            ws = writer.sheets["及格情况"]
            pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
            pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
            pie.height = 8
            pie.width = 12
            chart_sheet.add_chart(pie, "R1")
    return path
