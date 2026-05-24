from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.core.config import EXPORTS_DIR
from app.db.database import execute, fetch_all, fetch_one, get_connection, init_db


def _required(value: str | None, label: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{label}不能为空。")
    return text


def get_class_name(class_id: int) -> str:
    row = fetch_one("SELECT class_name, name FROM teaching_classes WHERE id = ?", (class_id,))
    if row is None:
        raise ValueError("未找到教学班。")
    return row.get("class_name") or row.get("name") or ""


def _next_anonymous_id(conn, table: str, where_sql: str, params: tuple) -> str:
    rows = conn.execute(
        f"""
        SELECT anonymous_id
        FROM {table}
        WHERE {where_sql}
        """,
        params,
    ).fetchall()
    max_index = 0
    for row in rows:
        text = str(row["anonymous_id"] or "").strip().upper()
        if text.startswith("S") and text[1:].isdigit():
            max_index = max(max_index, int(text[1:]))
    return f"S{max_index + 1:03d}"


def _normalize_student_no(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        head = text[:-2]
        if head.isdigit():
            return head
    return text


def _detect_columns(df: pd.DataFrame) -> tuple[str, str, str]:
    normalized = {str(column).strip().lower(): column for column in df.columns}
    no_candidates = ["学号", "student_no", "studentno", "学生学号"]
    name_candidates = ["姓名", "学生姓名", "name", "student_name", "studentname"]
    class_candidates = ["班级", "class", "class_name", "classname", "教学班"]

    def pick(candidates: list[str], label: str) -> str:
        for item in candidates:
            key = item.lower()
            if key in normalized:
                return normalized[key]
        raise ValueError(f"Excel 缺少必要列：{label}")

    return pick(no_candidates, "学号"), pick(name_candidates, "姓名"), pick(class_candidates, "班级")


def list_class_students(class_id: int) -> list[dict]:
    return fetch_all(
        """
        SELECT *
        FROM class_students
        WHERE class_id = ?
        ORDER BY anonymous_id ASC, student_no ASC, id ASC
        """,
        (class_id,),
    )


def list_course_students(course_id: int, class_id: int | None = None) -> list[dict]:
    params: list[int] = [course_id]
    where = "WHERE s.course_id = ?"
    if class_id is not None:
        where += " AND s.class_id = ?"
        params.append(class_id)
    return fetch_all(
        f"""
        SELECT s.*,
               c.course_name,
               tc.class_name AS teaching_class_name
        FROM students s
        LEFT JOIN courses c ON c.id = s.course_id
        LEFT JOIN teaching_classes tc ON tc.id = s.class_id
        {where}
        ORDER BY s.class_id ASC, s.anonymous_id ASC, s.student_no ASC, s.id ASC
        """,
        tuple(params),
    )


def create_class_student(class_id: int, student_no: str, student_name: str, class_name: str | None = None) -> int:
    student_no = _required(student_no, "学号")
    student_name = _required(student_name, "姓名")
    class_name = class_name or get_class_name(class_id)
    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id
            FROM class_students
            WHERE class_id = ? AND student_no = ?
            """,
            (class_id, student_no),
        ).fetchone()
        if existing:
            raise ValueError("当前班级基础名单中已存在该学号。")
        anonymous_id = _next_anonymous_id(conn, "class_students", "class_id = ?", (class_id,))
        cursor = conn.execute(
            """
            INSERT INTO class_students (class_id, student_no, name, student_name, class_name, anonymous_id)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (class_id, student_no, student_name, student_name, class_name, anonymous_id),
        )
        conn.commit()
        student_id = int(cursor.lastrowid)
    return student_id


def update_class_student(student_id: int, student_no: str, student_name: str, class_name: str | None = None) -> None:
    student_no = _required(student_no, "学号")
    student_name = _required(student_name, "姓名")
    row = fetch_one("SELECT * FROM class_students WHERE id = ?", (student_id,))
    if row is None:
        raise ValueError("未找到学生。")
    duplicate = fetch_one(
        """
        SELECT id
        FROM class_students
        WHERE class_id = ? AND student_no = ? AND id != ?
        """,
        (row.get("class_id"), student_no, student_id),
    )
    if duplicate:
        raise ValueError("当前班级基础名单中已存在该学号。")
    execute(
        """
        UPDATE class_students
        SET student_no = ?,
            name = ?,
            student_name = ?,
            class_name = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (student_no, student_name, student_name, class_name or row.get("class_name"), student_id),
    )


def delete_class_student(student_id: int) -> None:
    execute("DELETE FROM class_students WHERE id = ?", (student_id,))


def create_course_student(course_id: int, class_id: int, student_no: str, student_name: str, class_name: str | None = None) -> int:
    student_no = _required(student_no, "学号")
    student_name = _required(student_name, "姓名")
    class_name = class_name or get_class_name(class_id)
    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id
            FROM students
            WHERE course_id = ? AND class_id = ? AND student_no = ?
            """,
            (course_id, class_id, student_no),
        ).fetchone()
        if existing:
            raise ValueError("当前课程名单中已存在该学号。")
        anonymous_id = _next_anonymous_id(conn, "students", "course_id = ? AND class_id = ?", (course_id, class_id))
        cursor = conn.execute(
            """
            INSERT INTO students (course_id, class_id, student_no, name, student_name, class_name, anonymous_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (course_id, class_id, student_no, student_name, student_name, class_name, anonymous_id),
        )
        conn.commit()
        return int(cursor.lastrowid)


def update_course_student(student_id: int, student_no: str, student_name: str, class_name: str | None = None) -> None:
    student_no = _required(student_no, "学号")
    student_name = _required(student_name, "姓名")
    row = fetch_one("SELECT * FROM students WHERE id = ?", (student_id,))
    if row is None:
        raise ValueError("未找到课程名单学生。")
    duplicate = fetch_one(
        """
        SELECT id
        FROM students
        WHERE course_id = ? AND class_id = ? AND student_no = ? AND id != ?
        """,
        (row.get("course_id"), row.get("class_id"), student_no, student_id),
    )
    if duplicate:
        raise ValueError("当前课程名单副本中已存在该学号。")
    execute(
        """
        UPDATE students
        SET student_no = ?,
            name = ?,
            student_name = ?,
            class_name = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (student_no, student_name, student_name, class_name or row.get("class_name"), student_id),
    )


def delete_course_student(student_id: int) -> None:
    related = fetch_one("SELECT COUNT(*) AS count FROM submissions WHERE student_id = ?", (student_id,))
    if int((related or {}).get("count") or 0) > 0:
        raise ValueError("该学生已有实验报告提交或评分数据，不能从课程名单副本中删除。")
    execute("DELETE FROM students WHERE id = ?", (student_id,))


def sync_course_rosters_from_class(class_id: int) -> dict:
    init_db()
    with get_connection() as conn:
        course_rows = conn.execute(
            """
            SELECT course_id
            FROM course_class_links
            WHERE class_id = ?
              AND (deleted_at IS NULL OR TRIM(deleted_at) = '')
            ORDER BY course_id ASC
            """,
            (class_id,),
        ).fetchall()
    return sync_course_rosters_from_class_ids(class_id, [int(row["course_id"]) for row in course_rows])


def sync_course_roster_from_class(class_id: int, course_id: int) -> dict:
    return sync_course_rosters_from_class_ids(class_id, [course_id])


def sync_course_rosters_from_class_ids(class_id: int, course_ids: list[int]) -> dict:
    inserted = 0
    class_name = get_class_name(class_id)
    with get_connection() as conn:
        base_students = conn.execute(
            """
            SELECT *
            FROM class_students
            WHERE class_id = ?
            ORDER BY anonymous_id ASC, student_no ASC, id ASC
            """,
            (class_id,),
        ).fetchall()
        for course_id in sorted({int(course_id) for course_id in course_ids}):
            count = conn.execute(
                "SELECT COUNT(*) AS count FROM students WHERE course_id = ? AND class_id = ?",
                (course_id, class_id),
            ).fetchone()["count"]
            next_index = int(count or 0) + 1
            for student in base_students:
                existing = conn.execute(
                    """
                    SELECT id
                    FROM students
                    WHERE course_id = ? AND class_id = ? AND student_no = ?
                    """,
                    (course_id, class_id, student["student_no"]),
                ).fetchone()
                if existing:
                    continue
                anonymous_id = student["anonymous_id"] or f"S{next_index:03d}"
                conn.execute(
                    """
                    INSERT INTO students (course_id, class_id, student_no, name, student_name, class_name, anonymous_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        course_id,
                        class_id,
                        student["student_no"],
                        student["student_name"] or student["name"],
                        student["student_name"] or student["name"],
                        student["class_name"] or class_name,
                        anonymous_id,
                    ),
                )
                inserted += 1
                next_index += 1
        conn.commit()
    return {"inserted": inserted, "course_count": len(set(course_ids))}


def generate_student_template_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "学生名单"
    headers = ["学号", "姓名", "班级"]
    examples = [
        ["202201001", "张三", "22软工班"],
        ["202201002", "李四", "22软工班"],
        ["202201003", "王五", "22软工班"],
    ]
    ws.append(headers)
    for row in examples:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for index, width in enumerate([18, 14, 18], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"

    guide = wb.create_sheet("填写说明")
    guide_rows = [
        ["字段", "说明"],
        ["学号", "必填，建议使用文本格式，避免 Excel 自动改变长学号。"],
        ["姓名", "必填，请填写学生真实姓名。"],
        ["班级", "必填，可填写行政班、教学班或重修来源班级。"],
        ["表头", "不要修改第一行表头名称。"],
        ["重复", "同一教学班或同一课程名单副本内学号不能重复。"],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
    guide.column_dimensions["A"].width = 16
    guide.column_dimensions["B"].width = 70

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def import_class_students_from_excel(class_id: int, file_path: str | Path, mode: str = "append") -> dict:
    if mode not in {"append", "replace"}:
        raise ValueError("导入方式不正确。")
    df = pd.read_excel(file_path, dtype=str).fillna("")
    if df.empty:
        raise ValueError("Excel 中没有学生数据。")
    no_col, name_col, class_col = _detect_columns(df)

    rows: list[dict] = []
    seen: set[str] = set()
    duplicate_in_file: list[str] = []
    for _, row in df.iterrows():
        student_no = _normalize_student_no(row.get(no_col))
        student_name = str(row.get(name_col) or "").strip()
        class_name = str(row.get(class_col) or "").strip() or get_class_name(class_id)
        if not student_no or not student_name:
            continue
        if student_no in seen:
            duplicate_in_file.append(student_no)
            continue
        seen.add(student_no)
        rows.append({"student_no": student_no, "student_name": student_name, "class_name": class_name})

    if not rows:
        raise ValueError("Excel 中没有有效学生记录。")

    inserted = 0
    skipped_existing = 0
    replaced = 0
    with get_connection() as conn:
        if mode == "replace":
            replaced = int(conn.execute("SELECT COUNT(*) AS count FROM class_students WHERE class_id = ?", (class_id,)).fetchone()["count"] or 0)
            conn.execute("DELETE FROM class_students WHERE class_id = ?", (class_id,))
        for row in rows:
            existing = conn.execute(
                "SELECT id FROM class_students WHERE class_id = ? AND student_no = ?",
                (class_id, row["student_no"]),
            ).fetchone()
            if existing:
                skipped_existing += 1
                continue
            anonymous_id = _next_anonymous_id(conn, "class_students", "class_id = ?", (class_id,))
            conn.execute(
                """
                INSERT INTO class_students (class_id, student_no, name, student_name, class_name, anonymous_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (class_id, row["student_no"], row["student_name"], row["student_name"], row["class_name"], anonymous_id),
            )
            inserted += 1
        conn.commit()

    return {
        "inserted": inserted,
        "skipped_existing": skipped_existing,
        "duplicate_in_file": sorted(set(duplicate_in_file)),
        "replaced": replaced,
        "synced_course_students": 0,
    }


def _export_students_to_excel(rows: list[dict], filename_prefix: str) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    safe_prefix = "".join("_" if char in '\\/:*?"<>|' else char for char in filename_prefix).strip() or "学生名单"
    path = EXPORTS_DIR / f"{safe_prefix}.xlsx"
    export_rows = [
        {
            "学号": row.get("student_no") or "",
            "姓名": row.get("student_name") or row.get("name") or "",
            "班级": row.get("class_name") or row.get("teaching_class_name") or "",
            "匿名编号": row.get("anonymous_id") or "",
            "创建时间": row.get("created_at") or "",
        }
        for row in rows
    ]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name="学生名单")
        worksheet = writer.sheets["学生名单"]
        for width, column in zip([18, 14, 18, 12, 20], range(1, 6), strict=False):
            worksheet.column_dimensions[get_column_letter(column)].width = width
    return path


def export_class_students_to_excel(class_id: int) -> Path:
    class_name = get_class_name(class_id)
    rows = list_class_students(class_id)
    return _export_students_to_excel(rows, f"{class_name}_班级基础名单")


def export_course_students_to_excel(course_id: int, class_id: int | None = None) -> Path:
    course = fetch_one("SELECT course_name, name FROM courses WHERE id = ?", (course_id,))
    course_name = (course or {}).get("course_name") or (course or {}).get("name") or "课程"
    rows = list_course_students(course_id, class_id)
    suffix = "课程名单副本"
    if class_id:
        suffix = f"{get_class_name(class_id)}_课程名单副本"
    return _export_students_to_excel(rows, f"{course_name}_{suffix}")
